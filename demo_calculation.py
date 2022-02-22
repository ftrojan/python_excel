"""Demo of Python reading data from Excel, doing some calculation and saving it to another Excel."""
import logging
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font


def to_numpy(rng) -> np.ndarray:
    x = np.array([[c.value for c in r] for r in rng])
    return x


def write(x: np.ndarray, rng):
    for i, row in enumerate(rng):
        for j, cell in enumerate(row):
            # logging.debug(f"cell R{cell.row}C{cell.column} set to {x[i, j]}")
            cell.value = x[i, j]


def sumif_vec(a: np.ndarray, values: np.ndarray, unique_values: np.ndarray) -> np.ndarray:
    flat_values = np.ravel(values)
    result = np.array([
        np.sum(a[flat_values == value, :])
        for value in unique_values
    ]).reshape((-1, 1))
    return result


logging.basicConfig(level='INFO')
logging.debug("started")
input_filename = "sample_workbook.xlsx"
output_filename = "sample_output.xlsx"
logging.info(f"reading from {input_filename}")
wb = load_workbook(filename=input_filename, read_only=False)
ws = wb.active
logging.debug(ws.title)
numbers = to_numpy(ws['B3:B14'])
codes = to_numpy(ws["C3:C14"])
array = to_numpy(ws["D3:M14"])
unumbers = to_numpy(ws["B18:B20"])
ucodes = to_numpy(ws["B23:B26"])
logging.debug(f"number: {numbers.shape}")
logging.debug(f"code: {codes.shape}")
logging.debug(f"array: {array.shape}")
# calculation
total = np.sum(array, axis=1, keepdims=True)
total_by_number = sumif_vec(array, numbers, unumbers)
total_by_code = sumif_vec(array, codes, ucodes)
# write
write(total, ws["P3:P14"])
ws["P15"] = np.sum(total)
ws["P15"].font = Font(bold=True)
write(total_by_number, ws["P18:P20"])
ws["P21"] = np.sum(total_by_number)
ws["P21"].font = Font(bold=True)
write(total_by_code, ws["P23:P26"])
ws["P27"] = np.sum(total_by_code)
ws["P27"].font = Font(bold=True)
wb.save(filename=output_filename)
logging.info(f"output written to {output_filename}")
logging.debug("completed")
